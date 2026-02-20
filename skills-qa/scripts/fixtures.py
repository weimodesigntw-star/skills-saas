"""Shared test fixtures and helpers."""
import os, sys, time, json, shutil, subprocess, tempfile, tracemalloc
from pathlib import Path
from contextlib import contextmanager


def run_cmd(cmd, timeout=120):
    """Run subprocess, return (exit_code, stdout, stderr)."""
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    return r.returncode, r.stdout, r.stderr


def run_validator(script_path, *args, timeout=60):
    """Run a validation script, return (success, output)."""
    code, out, err = run_cmd([sys.executable, script_path, *args], timeout=timeout)
    return code == 0, (out + err).strip()


@contextmanager
def temp_dir():
    """Provide a temporary directory, cleaned up on exit."""
    d = tempfile.mkdtemp(prefix="skills_qa_")
    try:
        yield d
    finally:
        shutil.rmtree(d, ignore_errors=True)


@contextmanager
def measure_perf():
    """Context manager that measures time and memory. Yields a dict populated on exit."""
    metrics = {}
    tracemalloc.start()
    start = time.perf_counter()
    try:
        yield metrics
    finally:
        elapsed = time.perf_counter() - start
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        metrics["duration_sec"] = round(elapsed, 3)
        metrics["peak_memory_mb"] = round(peak / 1024 / 1024, 2)


def file_size_mb(path):
    return round(os.path.getsize(path) / 1024 / 1024, 3)


def create_test_xlsx(path, rows=10, cols=5, formulas=True):
    """Create a test XLSX with data and optional formulas."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestData"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=r * c)
    if formulas:
        from openpyxl.utils import get_column_letter
        fr = rows + 1
        for c in range(1, cols + 1):
            col = get_column_letter(c)
            ws.cell(row=fr, column=c, value=f"=SUM({col}1:{col}{rows})")
    wb.save(path)
    return path


def create_test_pdf(path, pages=3, text_per_page="Test content for page"):
    """Create a test PDF with reportlab."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    doc = SimpleDocTemplate(str(path), pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    for i in range(pages):
        story.append(Paragraph(f"{text_per_page} {i+1}", styles['Normal']))
        if i < pages - 1:
            story.append(PageBreak())
    doc.build(story)
    return path


def create_test_docx_js(path, content="Test Document"):
    """Create a DOCX using Node.js docx library."""
    js = f"""
const fs = require('fs');
const {{ Document, Packer, Paragraph, TextRun }} = require('docx');
const doc = new Document({{
  sections: [{{ children: [
    new Paragraph({{ children: [new TextRun({{ text: "{content}", bold: true, font: "Arial" }})] }})
  ] }}]
}});
Packer.toBuffer(doc).then(buf => {{ fs.writeFileSync("{path}", buf); process.exit(0); }});
"""
    code, _, err = run_cmd(["node", "-e", js], timeout=15)
    if code != 0:
        raise RuntimeError(f"docx-js failed: {err}")
    return path


def create_test_pptx_js(path, slides=3):
    """Create a PPTX using pptxgenjs."""
    slide_code = ""
    for i in range(slides):
        slide_code += f"""
let s{i} = pres.addSlide();
s{i}.addText('Slide {i+1}', {{ x: 1, y: 1, w: 8, h: 2, fontSize: 24 }});
"""
    js = f"""
const pptxgen = require('pptxgenjs');
const pres = new pptxgen();
pres.layout = 'LAYOUT_WIDE';
{slide_code}
pres.writeFile({{ fileName: '{path}' }}).then(() => process.exit(0));
"""
    code, _, err = run_cmd(["node", "-e", js], timeout=15)
    if code != 0:
        raise RuntimeError(f"pptxgenjs failed: {err}")
    return path
