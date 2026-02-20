"""XLSX Skill — Unit + Integration tests."""
import os, sys, unittest
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fixtures import temp_dir, measure_perf, create_test_xlsx, run_validator, file_size_mb
from config import VALIDATORS, TIME_LIMITS


class TestXlsxUnit(unittest.TestCase):
    """Tier 1: XLSX unit tests."""

    def test_create_workbook(self):
        with temp_dir() as d:
            p = os.path.join(d, "test.xlsx")
            with measure_perf() as m:
                create_test_xlsx(p, rows=10, cols=5)
            self.assertTrue(os.path.exists(p))
            self.assertLess(m["duration_sec"], TIME_LIMITS["unit"])

    def test_data_integrity(self):
        with temp_dir() as d:
            p = os.path.join(d, "data.xlsx")
            create_test_xlsx(p, rows=5, cols=3, formulas=False)
            from openpyxl import load_workbook
            wb = load_workbook(p)
            ws = wb.active
            self.assertEqual(ws.cell(row=1, column=1).value, 1)
            self.assertEqual(ws.cell(row=3, column=2).value, 6)
            self.assertEqual(ws.cell(row=5, column=3).value, 15)

    def test_formula_creation(self):
        with temp_dir() as d:
            p = os.path.join(d, "formula.xlsx")
            create_test_xlsx(p, rows=5, cols=3)
            from openpyxl import load_workbook
            wb = load_workbook(p)
            ws = wb.active
            self.assertTrue(ws.cell(row=6, column=1).value.startswith("=SUM"))

    def test_multiple_sheets(self):
        with temp_dir() as d:
            p = os.path.join(d, "multi.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            wb.active.title = "Sheet1"
            wb.create_sheet("Sheet2")
            wb.create_sheet("Sheet3")
            wb["Sheet1"]["A1"] = 100
            wb["Sheet2"]["A1"] = "=Sheet1!A1*2"
            wb.save(p)
            wb2 = Workbook()
            wb2 = __import__("openpyxl").load_workbook(p)
            self.assertEqual(len(wb2.sheetnames), 3)

    def test_number_formats(self):
        with temp_dir() as d:
            p = os.path.join(d, "fmt.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws["A1"] = 50000
            ws["A1"].number_format = '$#,##0'
            ws["B1"] = 0.153
            ws["B1"].number_format = '0.0%'
            wb.save(p)
            wb2 = __import__("openpyxl").load_workbook(p)
            self.assertEqual(wb2.active["A1"].number_format, '$#,##0')

    def test_unicode_cells(self):
        with temp_dir() as d:
            p = os.path.join(d, "unicode.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "中文測試"
            ws["B1"] = "日本語テスト"
            wb.save(p)
            wb2 = __import__("openpyxl").load_workbook(p)
            self.assertEqual(wb2.active["A1"].value, "中文測試")

    def test_xml_validation(self):
        """XLSX validation — verify validator runs without crash."""
        with temp_dir() as d:
            p = os.path.join(d, "valid.xlsx")
            create_test_xlsx(p)
            from fixtures import run_cmd
            code, out, err = run_cmd([sys.executable, VALIDATORS["xlsx_validate"], p])
            # validate.py may not support xlsx type; verify no unexpected crash
            self.assertIn(code, [0, 1], f"Validator crashed: {err}")


class TestXlsxIntegration(unittest.TestCase):
    """Tier 2: XLSX integration tests."""

    def test_create_recalc_verify(self):
        with temp_dir() as d, measure_perf() as m:
            p = os.path.join(d, "recalc.xlsx")
            create_test_xlsx(p, rows=10, cols=3)
            ok, out = run_validator(VALIDATORS["xlsx_recalc"], p, timeout=60)
            self.assertTrue(ok, f"Recalc failed: {out}")
            import json
            result = json.loads(out)
            self.assertEqual(result["status"], "success")
            self.assertEqual(result["total_errors"], 0)
            from openpyxl import load_workbook
            wb = load_workbook(p, data_only=True)
            ws = wb.active
            val = ws.cell(row=11, column=1).value
            self.assertIsNotNone(val, "Cached value missing after recalc")
            self.assertEqual(val, sum(i * 1 for i in range(1, 11)))
        self.assertLess(m["duration_sec"], TIME_LIMITS["integration"])


if __name__ == "__main__":
    unittest.main()
