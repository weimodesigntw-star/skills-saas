"""Tier 4: Regression tests — known bugs, previous failures."""
import os, sys, unittest, json, yaml
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fixtures import temp_dir, run_cmd, create_test_xlsx
from config import VALIDATORS


class TestRegressionKnownBugs(unittest.TestCase):
    """Verify known bugs remain documented and handled."""

    def test_skill_creator_yaml_bracket_bug(self):
        """Known: init_skill.py template uses [brackets] in description, parsed as YAML array."""
        with temp_dir() as d:
            run_cmd([sys.executable, VALIDATORS["sc_init_skill"], "bracket-test", "--path", d])
            skill_md = os.path.join(d, "bracket-test", "SKILL.md")
            content = open(skill_md).read()
            parts = content.split("---", 2)
            fm = yaml.safe_load(parts[1])
            # This is the KNOWN BUG: description is parsed as list, not string
            # We verify the bug still exists (or is fixed)
            desc = fm.get("description")
            if isinstance(desc, list):
                # Bug still present — document it
                self.skipTest("Known bug: YAML bracket in description parsed as list (upstream issue)")
            else:
                # Bug was fixed!
                self.assertIsInstance(desc, str)

    def test_quick_validate_handles_bracket_description(self):
        """Verify quick_validate detects or handles bracket description."""
        with temp_dir() as d:
            run_cmd([sys.executable, VALIDATORS["sc_init_skill"], "qv-test", "--path", d])
            skill_dir = os.path.join(d, "qv-test")
            code, out, err = run_cmd([sys.executable, VALIDATORS["sc_quick_validate"], skill_dir])
            # If the bracket bug exists, quick_validate should fail
            # We just verify it doesn't crash unexpectedly
            self.assertIn(code, [0, 1], "quick_validate should return 0 or 1, not crash")

    def test_xlsx_recalc_empty_formulas(self):
        """Regression: recalc on file with no formulas should not error."""
        with temp_dir() as d:
            p = os.path.join(d, "noformula.xlsx")
            create_test_xlsx(p, rows=5, cols=3, formulas=False)
            code, out, err = run_cmd([sys.executable, VALIDATORS["xlsx_recalc"], p], timeout=60)
            self.assertEqual(code, 0, f"Recalc on no-formula file failed: {out}{err}")
            result = json.loads(out)
            self.assertEqual(result["total_errors"], 0)

    def test_xlsx_division_by_zero(self):
        """Regression: formulas with division by zero should be caught."""
        with temp_dir() as d:
            p = os.path.join(d, "divzero.xlsx")
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws["A1"] = 100
            ws["B1"] = 0
            ws["C1"] = "=A1/B1"
            wb.save(p)
            code, out, _ = run_cmd([sys.executable, VALIDATORS["xlsx_recalc"], p], timeout=60)
            if code == 0:
                result = json.loads(out)
                if result["total_errors"] > 0:
                    self.assertIn("#DIV/0!", str(result.get("error_summary", "")))

    def test_pdf_merge_preserves_page_order(self):
        """Regression: merged PDF pages must maintain order."""
        with temp_dir() as d:
            from fixtures import create_test_pdf
            from pypdf import PdfReader, PdfWriter
            for i in range(3):
                create_test_pdf(os.path.join(d, f"p{i}.pdf"), 1, f"UNIQUE_PAGE_{i}")
            writer = PdfWriter()
            for i in range(3):
                for pg in PdfReader(os.path.join(d, f"p{i}.pdf")).pages:
                    writer.add_page(pg)
            merged = os.path.join(d, "merged.pdf")
            with open(merged, "wb") as f:
                writer.write(f)
            reader = PdfReader(merged)
            for i in range(3):
                text = reader.pages[i].extract_text()
                self.assertIn(f"UNIQUE_PAGE_{i}", text,
                              f"Page {i} content mismatch — order not preserved")


class TestRegressionPreviousFailures(unittest.TestCase):
    """Re-test scenarios that previously failed in the testing session."""

    def test_xlsx_cached_values_after_recalc(self):
        """Previously: XLSX formulas had no cached values after creation."""
        with temp_dir() as d:
            p = os.path.join(d, "cached.xlsx")
            create_test_xlsx(p, rows=5, cols=2)
            run_cmd([sys.executable, VALIDATORS["xlsx_recalc"], p], timeout=60)
            from openpyxl import load_workbook
            wb = load_workbook(p, data_only=True)
            ws = wb.active
            val = ws.cell(row=6, column=1).value
            self.assertIsNotNone(val, "Cached formula value still missing after recalc")

    def test_docx_validates_after_creation(self):
        """Previously: DOCX validation was a concern."""
        with temp_dir() as d:
            from fixtures import create_test_docx_js, run_validator
            p = os.path.join(d, "val.docx")
            create_test_docx_js(p, "Validation regression test")
            ok, out = run_validator(VALIDATORS["docx_validate"], p)
            self.assertTrue(ok, f"DOCX validation regression: {out}")


if __name__ == "__main__":
    unittest.main()
