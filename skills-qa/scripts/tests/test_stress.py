"""Tier 3: Stress tests — large files, concurrency, memory, edge cases."""
import os, sys, gc, unittest, tracemalloc, time
from concurrent.futures import ProcessPoolExecutor, as_completed
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from fixtures import temp_dir, measure_perf, create_test_xlsx, create_test_pdf, file_size_mb
from config import STRESS, TIME_LIMITS, MEMORY_LIMITS


# ── Worker functions (must be top-level for ProcessPoolExecutor) ──

def _worker_create_xlsx(path):
    from fixtures import create_test_xlsx
    create_test_xlsx(path, rows=100, cols=10)
    return os.path.getsize(path)

def _worker_create_pdf(path):
    from fixtures import create_test_pdf
    create_test_pdf(path, pages=10)
    return os.path.getsize(path)


class TestStressLargeFiles(unittest.TestCase):
    """Stress: large file generation."""

    def test_xlsx_10k_rows(self):
        with temp_dir() as d, measure_perf() as m:
            p = os.path.join(d, "big.xlsx")
            create_test_xlsx(p, rows=STRESS["xlsx_rows"], cols=STRESS["xlsx_cols"], formulas=True)
            self.assertTrue(os.path.exists(p))
            self.assertGreater(os.path.getsize(p), 0)
        self.assertLess(m["duration_sec"], TIME_LIMITS["stress_large_file"],
                        f"XLSX 10K rows took {m['duration_sec']}s")
        self.assertLess(m["peak_memory_mb"], MEMORY_LIMITS["stress"],
                        f"Peak memory {m['peak_memory_mb']}MB")

    def test_pdf_100_pages(self):
        with temp_dir() as d, measure_perf() as m:
            p = os.path.join(d, "big.pdf")
            create_test_pdf(p, pages=STRESS["pdf_pages"], text_per_page="Stress test content " * 50)
            self.assertTrue(os.path.exists(p))
            from pypdf import PdfReader
            self.assertEqual(len(PdfReader(p).pages), STRESS["pdf_pages"])
        self.assertLess(m["duration_sec"], TIME_LIMITS["stress_large_file"])

    def test_xlsx_formula_stress(self):
        """10K rows with formulas, then recalculate."""
        with temp_dir() as d, measure_perf() as m:
            p = os.path.join(d, "formulas.xlsx")
            create_test_xlsx(p, rows=1000, cols=10, formulas=True)
            from fixtures import run_validator
            from config import VALIDATORS
            ok, out = run_validator(VALIDATORS["xlsx_recalc"], p, timeout=90)
            self.assertTrue(ok, f"Recalc stress failed: {out}")
        self.assertLess(m["duration_sec"], TIME_LIMITS["stress_large_file"])


class TestStressConcurrency(unittest.TestCase):
    """Stress: concurrent operations."""

    def test_parallel_xlsx_creation(self):
        with temp_dir() as d, measure_perf() as m:
            paths = [os.path.join(d, f"file_{i}.xlsx") for i in range(STRESS["concurrent_workers"])]
            with ProcessPoolExecutor(max_workers=STRESS["concurrent_workers"]) as executor:
                futures = {executor.submit(_worker_create_xlsx, p): p for p in paths}
                for f in as_completed(futures):
                    size = f.result()
                    self.assertGreater(size, 0)
            for p in paths:
                self.assertTrue(os.path.exists(p))
        self.assertLess(m["duration_sec"], TIME_LIMITS["stress_concurrent"])

    def test_parallel_pdf_creation(self):
        with temp_dir() as d, measure_perf() as m:
            paths = [os.path.join(d, f"file_{i}.pdf") for i in range(STRESS["concurrent_workers"])]
            with ProcessPoolExecutor(max_workers=STRESS["concurrent_workers"]) as executor:
                futures = {executor.submit(_worker_create_pdf, p): p for p in paths}
                for f in as_completed(futures):
                    size = f.result()
                    self.assertGreater(size, 0)
        self.assertLess(m["duration_sec"], TIME_LIMITS["stress_concurrent"])


class TestStressMemory(unittest.TestCase):
    """Stress: memory leak detection."""

    def test_no_memory_leak_xlsx(self):
        """Create files in a loop, verify memory returns to baseline."""
        gc.collect()
        tracemalloc.start()
        baseline = tracemalloc.get_traced_memory()[0]
        with temp_dir() as d:
            for i in range(STRESS["memory_leak_iterations"]):
                p = os.path.join(d, f"leak_{i}.xlsx")
                create_test_xlsx(p, rows=100, cols=10)
                os.remove(p)
                gc.collect()
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        growth_mb = (current - baseline) / 1024 / 1024
        self.assertLess(growth_mb, 50, f"Memory grew {growth_mb:.1f}MB over {STRESS['memory_leak_iterations']} iterations")

    def test_no_memory_leak_pdf(self):
        gc.collect()
        tracemalloc.start()
        baseline = tracemalloc.get_traced_memory()[0]
        with temp_dir() as d:
            for i in range(STRESS["memory_leak_iterations"]):
                p = os.path.join(d, f"leak_{i}.pdf")
                create_test_pdf(p, pages=5)
                os.remove(p)
                gc.collect()
        current, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        growth_mb = (current - baseline) / 1024 / 1024
        self.assertLess(growth_mb, 50, f"Memory grew {growth_mb:.1f}MB")


class TestStressEdgeCases(unittest.TestCase):
    """Stress: edge cases and boundary conditions."""

    def test_unicode_all_scripts(self):
        """Test multiple scripts: CJK, Arabic, Cyrillic."""
        with temp_dir() as d:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            texts = ["English", "中文繁體", "العربية", "Кириллица", "ひらがな", "한국어"]
            for i, t in enumerate(texts, 1):
                ws.cell(row=i, column=1, value=t)
            p = os.path.join(d, "unicode.xlsx")
            wb.save(p)
            wb2 = __import__("openpyxl").load_workbook(p)
            for i, t in enumerate(texts, 1):
                self.assertEqual(wb2.active.cell(row=i, column=1).value, t)

    def test_empty_xlsx(self):
        with temp_dir() as d:
            from openpyxl import Workbook
            p = os.path.join(d, "empty.xlsx")
            Workbook().save(p)
            self.assertTrue(os.path.exists(p))
            self.assertGreater(os.path.getsize(p), 0)

    def test_empty_pdf(self):
        with temp_dir() as d:
            p = os.path.join(d, "empty.pdf")
            create_test_pdf(p, pages=1, text_per_page=" ")
            from pypdf import PdfReader
            self.assertEqual(len(PdfReader(p).pages), 1)

    def test_long_cell_content(self):
        """Test Excel cell with ~30K characters."""
        with temp_dir() as d:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            long_text = "A" * 30000
            ws["A1"] = long_text
            p = os.path.join(d, "long.xlsx")
            wb.save(p)
            wb2 = __import__("openpyxl").load_workbook(p)
            self.assertEqual(len(wb2.active["A1"].value), 30000)

    def test_special_chars_in_xlsx(self):
        """Test special characters: quotes, brackets, ampersands."""
        with temp_dir() as d:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            specials = ['He said "hello"', "It's <here> & there", "a\tb\nc", "[brackets]"]
            for i, s in enumerate(specials, 1):
                ws.cell(row=i, column=1, value=s)
            p = os.path.join(d, "special.xlsx")
            wb.save(p)
            wb2 = __import__("openpyxl").load_workbook(p)
            for i, s in enumerate(specials, 1):
                self.assertEqual(wb2.active.cell(row=i, column=1).value, s)


if __name__ == "__main__":
    unittest.main()
