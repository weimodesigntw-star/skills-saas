"""Report generator — JSON + XLSX + Console output."""
import json, datetime


def generate_json_report(results, total_time):
    """Build structured JSON report from test results."""
    passed = sum(1 for r in results if r["status"] == "passed")
    failed = sum(1 for r in results if r["status"] == "failed")
    errors = sum(1 for r in results if r["status"] == "error")
    skipped = sum(1 for r in results if r["status"] == "skipped")
    total = len(results)

    # Group by tier
    by_tier = {}
    for r in results:
        t = f"tier_{r['tier']}"
        if t not in by_tier:
            by_tier[t] = {"passed": 0, "failed": 0, "errors": 0, "skipped": 0, "total": 0}
        by_tier[t]["total"] += 1
        if r["status"] == "passed":
            by_tier[t]["passed"] += 1
        elif r["status"] == "failed":
            by_tier[t]["failed"] += 1
        elif r["status"] == "error":
            by_tier[t]["errors"] += 1
        elif r["status"] == "skipped":
            by_tier[t]["skipped"] += 1

    # Group by skill
    by_skill = {}
    for r in results:
        s = r["skill"]
        if s not in by_skill:
            by_skill[s] = {"passed": 0, "failed": 0, "errors": 0, "skipped": 0, "total": 0}
        by_skill[s]["total"] += 1
        if r["status"] == "passed":
            by_skill[s]["passed"] += 1
        elif r["status"] == "failed":
            by_skill[s]["failed"] += 1
        elif r["status"] == "error":
            by_skill[s]["errors"] += 1
        elif r["status"] == "skipped":
            by_skill[s]["skipped"] += 1

    return {
        "metadata": {
            "timestamp": datetime.datetime.now().isoformat(),
            "total_duration_sec": round(total_time, 2),
            "framework": "unittest",
        },
        "summary": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "errors": errors,
            "skipped": skipped,
            "pass_rate": round(passed / total, 4) if total > 0 else 0,
        },
        "by_tier": by_tier,
        "by_skill": by_skill,
        "tests": results,
        "slowest_tests": sorted(results, key=lambda x: x["duration_sec"], reverse=True)[:10],
    }


def generate_xlsx_report(report, path):
    """Generate Excel summary report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    header_fill = PatternFill("solid", fgColor="1E2761")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    pass_fill = PatternFill("solid", fgColor="D5F5E3")
    fail_fill = PatternFill("solid", fgColor="FADBD8")
    skip_fill = PatternFill("solid", fgColor="FFF3CD")

    def write_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    s = report["summary"]
    ws["A1"] = "Skills QA Test Report"
    ws["A1"].font = Font(bold=True, size=16, name="Arial", color="1E2761")
    ws["A2"] = f"Generated: {report['metadata']['timestamp']}"
    ws["A3"] = f"Duration: {report['metadata']['total_duration_sec']}s"

    write_header(ws, ["Metric", "Value"], row=5)
    metrics = [("Total Tests", s["total"]), ("Passed", s["passed"]), ("Failed", s["failed"]),
               ("Errors", s["errors"]), ("Skipped", s["skipped"]),
               ("Pass Rate", f"{s['pass_rate']*100:.1f}%")]
    for i, (k, v) in enumerate(metrics, 6):
        ws.cell(row=i, column=1, value=k).border = border
        c = ws.cell(row=i, column=2, value=v)
        c.border = border
        c.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # ── Sheet 2: By Skill ──
    ws2 = wb.create_sheet("By Skill")
    write_header(ws2, ["Skill", "Total", "Passed", "Failed", "Errors", "Skipped", "Pass Rate"])
    row = 2
    for skill, data in sorted(report["by_skill"].items()):
        ws2.cell(row=row, column=1, value=skill).border = border
        ws2.cell(row=row, column=2, value=data["total"]).border = border
        ws2.cell(row=row, column=3, value=data["passed"]).border = border
        ws2.cell(row=row, column=4, value=data["failed"]).border = border
        ws2.cell(row=row, column=5, value=data["errors"]).border = border
        ws2.cell(row=row, column=6, value=data["skipped"]).border = border
        rate = data["passed"] / data["total"] if data["total"] > 0 else 0
        c = ws2.cell(row=row, column=7, value=f"{rate*100:.0f}%")
        c.border = border
        c.fill = pass_fill if rate == 1 else (fail_fill if rate < 0.7 else skip_fill)
        row += 1
    for col in 'ABCDEFG':
        ws2.column_dimensions[col].width = 14

    # ── Sheet 3: All Tests ──
    ws3 = wb.create_sheet("All Tests")
    write_header(ws3, ["Test", "Tier", "Skill", "Status", "Duration (s)", "Error"])
    for i, t in enumerate(report["tests"], 2):
        ws3.cell(row=i, column=1, value=t["name"]).border = border
        ws3.cell(row=i, column=2, value=t["tier"]).border = border
        ws3.cell(row=i, column=3, value=t["skill"]).border = border
        c = ws3.cell(row=i, column=4, value=t["status"])
        c.border = border
        c.fill = pass_fill if t["status"] == "passed" else (
            fail_fill if t["status"] in ("failed", "error") else skip_fill)
        ws3.cell(row=i, column=5, value=t["duration_sec"]).border = border
        ws3.cell(row=i, column=6, value=(t.get("error") or "")[:100]).border = border
    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['F'].width = 50

    # ── Sheet 4: Slowest Tests ──
    ws4 = wb.create_sheet("Slowest Tests")
    write_header(ws4, ["Rank", "Test", "Duration (s)", "Skill"])
    for i, t in enumerate(report["slowest_tests"], 2):
        ws4.cell(row=i, column=1, value=i - 1).border = border
        ws4.cell(row=i, column=2, value=t["name"]).border = border
        ws4.cell(row=i, column=3, value=t["duration_sec"]).border = border
        ws4.cell(row=i, column=4, value=t["skill"]).border = border
    ws4.column_dimensions['B'].width = 40

    wb.save(path)


def print_console_summary(report):
    """Print colored console summary."""
    s = report["summary"]
    print("\n" + "=" * 60)
    print("  TEST RESULTS")
    print("=" * 60)

    # Per-skill summary
    for skill, data in sorted(report["by_skill"].items()):
        rate = data["passed"] / data["total"] if data["total"] > 0 else 0
        icon = "PASS" if rate == 1 else ("PARTIAL" if rate >= 0.5 else "FAIL")
        bar = "#" * int(rate * 20) + "-" * (20 - int(rate * 20))
        print(f"  {skill:<18} [{bar}] {data['passed']}/{data['total']} {icon}")

    print()
    print(f"  Total: {s['total']} | Passed: {s['passed']} | Failed: {s['failed']} | "
          f"Errors: {s['errors']} | Skipped: {s['skipped']}")
    print(f"  Pass Rate: {s['pass_rate']*100:.1f}%")
    print(f"  Duration: {report['metadata']['total_duration_sec']}s")

    # Show failures
    failures = [t for t in report["tests"] if t["status"] in ("failed", "error")]
    if failures:
        print(f"\n  FAILURES ({len(failures)}):")
        for t in failures:
            print(f"    FAIL {t['name']}")
            if t.get("error"):
                err_short = t["error"].split("\n")[0][:80]
                print(f"         {err_short}")

    print("=" * 60)
